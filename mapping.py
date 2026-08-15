"""
Maps a veromass-aligner `aligned_features.xlsx` workbook into the commit
body shapes expected by moleculeid-api's POST /api/jobs/{job_id}/commit
(see moleculeid-api/workbench_routes.py: FeatureInput, JobCommit).

Targeted jobs  -> {"features": [{"mz", "rt", "intensity", "fragments"}, ...]}
Untargeted jobs -> {"feature_matrix": {"<feature>": {"<sample>": <intensity>}}}

Pure functions, no network/auth here — easy to unit-test against a real
workbook produced by veromass-aligner.
"""

import os
import re

import openpyxl


def _read_sheet_as_dicts(ws):
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    return [dict(zip(header, row)) for row in rows]


_TOTAL_FILES_RE = re.compile(r"Total MS files:\s*(\d+)")


def _expected_sample_count(xlsx_path):
    """Read the aligner's own `alignment_log.txt` (written alongside
    aligned_features.xlsx in the same output folder) for its "Total MS
    files: N" line — an INDEPENDENT count of how many samples this run
    actually processed, logged at the very start of the run, before any
    grouping/correspondence/Excel-writing step could lose one. Returns None
    (not an error) when the log is missing or the line can't be found —
    e.g. an older aligner output, or the log was moved/deleted — since the
    whole point is a sanity check on data we ALREADY have, not a new hard
    requirement that makes otherwise-valid workbooks unusable."""
    log_path = os.path.join(os.path.dirname(xlsx_path), "alignment_log.txt")
    if not os.path.isfile(log_path):
        return None
    try:
        with open(log_path, "r", encoding="utf-8", errors="ignore") as fh:
            text = fh.read()
    except OSError:
        return None
    m = _TOTAL_FILES_RE.search(text)
    return int(m.group(1)) if m else None


def _check_sample_count(xlsx_path, feature_matrix):
    """Cross-validate the feature_matrix this run is about to commit against
    the aligner's own independently-logged file count. Found live: a job
    committed with one fewer sample than the aligner actually processed,
    root-caused to a stale output folder from a prior run (fixed at the
    source in VeroMass_Aligner v1.11.3 — see its "Cleared stale output..."
    log line) — this is the second, independent layer: even if a future
    bug reintroduces sample loss through some OTHER path, this catches it
    HERE, before any network call, rather than silently committing a
    truncated matrix the way the original incident did."""
    if not feature_matrix:
        return
    expected = _expected_sample_count(xlsx_path)
    if expected is None:
        return
    actual = len({s for row in feature_matrix.values() for s in row})
    if actual != expected:
        raise ValueError(
            f"Sample count mismatch: aligned_features.xlsx has {actual} sample(s) in its "
            f"feature matrix, but alignment_log.txt reports {expected} file(s) were processed. "
            "This usually means the output folder had leftover data from a previous run — "
            "re-run the alignment into a clean folder and try again."
        )


def build_targeted_features(xlsx_path):
    """Read the Features sheet -> list of FeatureInput-shaped dicts.

    intensity uses Base.Peak (the representative peak height for the
    feature); fragments is passed through as-is since MS2.Fragments is
    already in the server's expected "mz(pct%); mz(pct%)" string format.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Features"]
    out = []
    for row in _read_sheet_as_dicts(ws):
        mz = row.get("m.z")
        rt = row.get("RT")
        if mz is None or rt is None:
            continue
        out.append({
            "mz": float(mz),
            "rt": float(rt),
            "intensity": float(row.get("Base.Peak") or 0.0),
            "fragments": row.get("MS2.Fragments") or None,
        })
    return out


def build_untargeted_feature_matrix(xlsx_path):
    """Read the Intensities sheet -> {feature: {sample: intensity}}."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Intensities"]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    feature_col, sample_cols = header[0], header[1:]

    matrix = {}
    for row in rows:
        row = dict(zip(header, row))
        feature = row.get(feature_col)
        if feature is None:
            continue
        matrix[str(feature)] = {
            sample: float(row[sample]) if row.get(sample) is not None else 0.0
            for sample in sample_cols
        }
    return matrix


def build_untargeted_feature_meta(xlsx_path):
    """Read the Features sheet -> {feature: {mz, rt}} — this is what makes
    POST /api/jobs/{id}/annotate possible server-side (see
    workbench_routes.py's JobCommit.feature_meta docstring): without it, an
    untargeted job's "compounds" are just row labels with no coordinates to
    match against the library. Feature ids here (Feature_000001, ...) are the
    SAME ones write_feature_table used across Features/Peaks/Intensities, so
    they line up with build_untargeted_feature_matrix's keys directly."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Features"]
    meta = {}
    for i, row in enumerate(_read_sheet_as_dicts(ws), 1):
        feature_id = f"Feature_{i:06d}"
        mz, rt = row.get("m.z"), row.get("RT")
        if mz is None or rt is None:
            continue
        meta[feature_id] = {"mz": float(mz), "rt": float(rt)}
    return meta


def build_chromatograms(xlsx_path):
    """Read the optional TIC / BPI / RT_Correction sheets (absent on older
    aligner output, or on an all-MGF run with no scan-level data — see
    VeroMass_Aligner.py's write_feature_table docstring) into the per-sample
    trace shape Workbench's chart components expect:
      {"tic": [{"sample","x","y"}, ...], "bpi": [...],
       "rt_correction": [{"sample","rtRaw","rtDeviation"}, ...]}
    Returns an empty dict (not None) when none of the sheets are present, so
    callers can always do `commit_body.get("chromatograms") or {}` safely."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    result = {}

    for sheet, key, y_field in (("TIC", "tic", "intensity"), ("BPI", "bpi", "intensity")):
        if sheet not in wb.sheetnames:
            continue
        by_sample = {}
        for row in _read_sheet_as_dicts(wb[sheet]):
            s = row.get("sample")
            if s is None or row.get("rt_min") is None:
                continue
            by_sample.setdefault(s, {"x": [], "y": []})
            by_sample[s]["x"].append(float(row["rt_min"]))
            by_sample[s]["y"].append(float(row[y_field] or 0.0))
        result[key] = [{"sample": s, **traces} for s, traces in by_sample.items()]

    if "RT_Correction" in wb.sheetnames:
        by_sample = {}
        for row in _read_sheet_as_dicts(wb["RT_Correction"]):
            s = row.get("sample")
            if s is None or row.get("rt_raw_min") is None:
                continue
            by_sample.setdefault(s, {"rtRaw": [], "rtDeviation": []})
            by_sample[s]["rtRaw"].append(float(row["rt_raw_min"]))
            by_sample[s]["rtDeviation"].append(float(row.get("rt_deviation_min") or 0.0))
        result["rt_correction"] = [{"sample": s, **curve} for s, curve in by_sample.items()]

    return result


def build_commit_payload(xlsx_path, mode):
    """mode: "targeted" or "untargeted" -> the mode-specific body fields
    (package_uuid is added by the caller, not here)."""
    if mode == "targeted":
        return {"features": build_targeted_features(xlsx_path)}
    if mode == "untargeted":
        feature_matrix = build_untargeted_feature_matrix(xlsx_path)
        _check_sample_count(xlsx_path, feature_matrix)
        payload = {
            "feature_matrix": feature_matrix,
            "feature_meta": build_untargeted_feature_meta(xlsx_path),
        }
        chromatograms = build_chromatograms(xlsx_path)
        if chromatograms:
            payload["chromatograms"] = chromatograms
        return payload
    raise ValueError(f"Unknown job mode: {mode!r}")
